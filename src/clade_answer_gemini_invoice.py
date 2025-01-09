import os
import json
from dotenv import load_dotenv
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, PatternFill
import re
from fnmatch import fnmatch
import pdf2image
from PIL import Image
import xlsxwriter
import shutil
from typing import Dict, Any, List, Optional, Set, Literal
from enum import Enum
import traceback
from datetime import datetime
import uuid

from app.tools.doc_proccesing.claude_vertex_tool import ClaudeVertexDocumentTool
from app.tools.doc_proccesing.gemini_vertex_tool import GeminiVertexDocumentTool  


class VerificationType(Enum):
    CLAUDE = "claude"
    GEMINI = "gemini"
    SKIP = "skip"

class PromptLoader:
    """Handles loading and managing AI prompts from text files."""
    
    def __init__(self, prompts_dir: str = "prompts"):
        self.prompts_dir = Path(prompts_dir)
        if not self.prompts_dir.exists():
            self.prompts_dir.mkdir(parents=True)

    def load_prompt(self, filename: str) -> Dict[str, str]:
        """Load prompt from text file and parse sections."""
        try:
            file_path = self.prompts_dir / filename
            if not file_path.exists():
                print(f"Prompt file {filename} not found, using default prompt.")
                return self._get_default_prompt(filename)
            
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()

            # Extract system prompt and main prompt
            system_prompt = self._extract_section(content, "system")
            prompt_template = self._extract_section(content, "prompt")

            return {
                "system_prompt": system_prompt.strip(),
                "question_template": prompt_template.strip()
            }

        except Exception as e:
            print(f"Error loading prompt file {filename}: {str(e)}")
            return self._get_default_prompt(filename)

    def _extract_section(self, content: str, section_name: str) -> str:
        """Extract content between section tags."""
        try:
            start_tag = f"<{section_name}>"
            end_tag = f"</{section_name}>"
            
            start = content.find(start_tag) + len(start_tag)
            end = content.find(end_tag)
            
            if start == -1 or end == -1:
                raise ValueError(f"Could not find {section_name} section in prompt file")
                
            return content[start:end].strip()
        except Exception as e:
            print(f"Error extracting {section_name} section: {str(e)}")
            raise

    def _get_default_prompt(self, filename: str) -> Dict[str, str]:
        """Return default prompt based on filename."""
        if "claude" in filename.lower():
            return {
                "system_prompt": "You are an invoice verification expert. Extract and validate information according to the provided structure.",
                "question_template": """Examine this invoice and verify all fields.
Reference data: {json_data}
Return only a valid JSON object matching the reference structure."""
            }
        else:
            return {
                "system_prompt": "You are an invoice analysis expert.",
                "question_template": """Analyze this invoice data.
Input: {json_data}
Return only a valid JSON object."""
            }

    def format_prompt(self, prompt_data: Dict[str, str], **kwargs) -> Dict[str, str]:
        """Format prompt template with provided variables."""
        try:
            # Pre-process JSON data if present
            if 'json_data' in kwargs:
                if isinstance(kwargs['json_data'], (dict, list)):
                    kwargs['json_data'] = json.dumps(kwargs['json_data'], indent=2, ensure_ascii=False)
            
            # Format the prompt template
            formatted_question = prompt_data["question_template"]
            for key, value in kwargs.items():
                placeholder = '{' + key + '}'
                formatted_question = formatted_question.replace(placeholder, str(value))
            
            return {
                "system_prompt": prompt_data["system_prompt"],
                "question": formatted_question
            }
        except Exception as e:
            print(f"Error formatting prompt: {str(e)}")
            return prompt_data




class DocumentProcessor:
    """Handles processing of individual documents."""
    
    def __init__(self, excel_file: str, images_dir: Path, ignored_patterns: Optional[Set[str]] = None):
        self.excel_file = excel_file
        self.images_dir = images_dir
        self.claude_tool = ClaudeVertexDocumentTool()
        self.gemini_tool = GeminiVertexDocumentTool()
        self.prompt_loader = PromptLoader()
        self.ignored_patterns = ignored_patterns if ignored_patterns is not None else set()

    def verify_with_claude(self, json_file: Path, pdf_file: Path) -> Dict:
        """Get verification from Claude."""
        try:
            print(f"\nStarting Claude verification for {json_file.name}")
            
            # Load JSON data
            with open(json_file, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            
            # Load and format Claude prompt with .txt extension
            prompt_config = self.prompt_loader.load_prompt("claude_prompt.txt")
            formatted_prompt = self.prompt_loader.format_prompt(
                prompt_config,
                json_data=json.dumps(json_data, indent=2)
            )
            
            # Get verification from Claude
            result = self.claude_tool.analyze_documents(
                file_locations=[str(json_file), str(pdf_file)],
                system_prompt=formatted_prompt["system_prompt"],
                question=formatted_prompt["question"],
                return_json_only=True
            )
            
            return self._parse_ai_response(result)

        except Exception as e:
            print(f"Error in Claude verification: {str(e)}")
            return {}

    def verify_with_gemini(self, json_file: Path, pdf_file: Path) -> Dict:
        """Get verification from Gemini."""
        try:
            print(f"\nStarting Gemini verification for {json_file.name}")
            
            # Load PDF content
            with open(pdf_file, 'rb') as f:
                pdf_content = f.read()
            
            # Load and format Gemini prompt with .txt extension
            prompt_config = self.prompt_loader.load_prompt("gemini_prompt.txt")
            formatted_prompt = self.prompt_loader.format_prompt(
                prompt_config,
                pdf_content=pdf_content
            )
            
            # Get verification from Gemini
            result = self.gemini_tool.verify_invoice(
                prompt=formatted_prompt["question"],
                system_prompt=formatted_prompt["system_prompt"],
                pdf_content=pdf_content
            )
            
            return self._parse_ai_response(result)

        except Exception as e:
            print(f"Error in Gemini verification: {str(e)}")
            return {}
        
    def update_sheet_with_verification(self, sheet_name: str, ai_data: Dict, verifier: str):
        """Update sheet with verification results and colored match status."""
        print(f"\nUpdating sheet {sheet_name} with {verifier} verification")
        
        wb = openpyxl.load_workbook(self.excel_file)
        sheet = wb[sheet_name]
        
        flattened_ai = self._flatten_json(ai_data)
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Define colors for match status
        green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green
        red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')    # Light red
        
        # Define font colors
        from openpyxl.styles import Font
        green_font = Font(color='006400')  # Dark green
        red_font = Font(color='8B0000')    # Dark red
        
        for row in range(2, sheet.max_row + 1):
            field = sheet.cell(row=row, column=1).value
            if field:
                gemini_value = sheet.cell(row=row, column=2).value
                ai_value = str(flattened_ai.get(field, ''))
                
                # Update AI value (Claude or Gemini)
                value_column = 3  # Claude column
                if verifier.lower() == 'gemini':
                    value_column = 2  # Gemini column
                
                sheet.cell(row=row, column=value_column, value=ai_value)
                
                # Update match status with color
                if gemini_value and ai_value:
                    match = 'Yes' if str(gemini_value).strip() == ai_value.strip() else 'No'
                    match_cell = sheet.cell(row=row, column=4, value=match)
                    
                    # Apply background and font color based on match status
                    if match == 'Yes':
                        match_cell.fill = green_fill
                        match_cell.font = green_font
                    else:
                        match_cell.fill = red_fill
                        match_cell.font = red_font
                
                # Update timestamp and verifier
                sheet.cell(row=row, column=5, value=timestamp)
                sheet.cell(row=row, column=6, value=verifier)
        
        wb.save(self.excel_file)
        print(f"Updated sheet with colored match status (both background and text)")
        

    def _clean_sheet_name(self, name: str) -> str:
        """Clean filename to make it a valid Excel sheet name."""
        try:
            # Remove file extension
            name = os.path.splitext(name)[0]
            
            # Replace invalid characters with underscore
            name = re.sub(r'[\\/*?[\]:]+', '_', name)
            
            # Handle sheet name length limit (31 chars)
            if len(name) > 31:
                # For invoice numbers, try to preserve the numbers and important parts
                parts = name.split('_')
                if len(parts) >= 2 and parts[0].isdigit() and parts[1].isdigit():
                    # Keep the invoice numbers and add a shortened version of the rest
                    invoice_nums = f"{parts[0]}_{parts[1]}"
                    remaining = '_'.join(parts[2:])
                    
                    # Calculate remaining space
                    remaining_space = 31 - len(invoice_nums) - 1  # -1 for underscore
                    
                    if remaining_space > 0:
                        # Take first few chars of remaining parts
                        shortened = remaining[:remaining_space]
                        name = f"{invoice_nums}_{shortened}"
                    else:
                        name = invoice_nums
                else:
                    # If not invoice numbers, take start and end
                    name = name[:15] + '_' + name[-15:]
            
            print(f"Sheet name cleaning: {name}")
            return name
            
        except Exception as e:
            print(f"Error cleaning sheet name: {str(e)}")
            return "Sheet_" + str(uuid.uuid4())[:8]


    def add_image_to_sheet(self, sheet_name: str, pdf_file: Path):
        """Convert PDF and add image to sheet while preserving existing sheets and images."""
        print(f"\nAdding image to sheet {sheet_name}")
        
        temp_excel = f"temp_{uuid.uuid4()}_{self.excel_file}"
        
        try:
            # Convert PDF to image
            images = pdf2image.convert_from_path(
                pdf_file,
                dpi=150,
                first_page=1,
                last_page=1,
                size=(800, None),
                fmt="png",
                thread_count=2
            )
            
            if not images:
                print(f"No images converted from PDF: {pdf_file}")
                return
                
            # Save the image
            img_path = self.images_dir / f"{pdf_file.stem}.png"
            images[0].save(str(img_path), "PNG", optimize=True)
            
            print(f"Image saved successfully at {img_path}")
            
            # Store the current sheet's image locations
            existing_images = {}
            try:
                source_wb = openpyxl.load_workbook(self.excel_file)
                # Store any existing image references
                for sname in source_wb.sheetnames:
                    if sname != sheet_name:  # Only store other sheets' images
                        sheet = source_wb[sname]
                        img_cell = sheet['H1'].value
                        if img_cell and img_cell.startswith('Image:'):
                            existing_images[sname] = img_cell.replace('Image:', '').strip()
                source_wb.close()
            except Exception as e:
                print(f"Warning when reading existing images: {str(e)}")
            
            # Create new workbook
            workbook = xlsxwriter.Workbook(temp_excel)
            
            try:
                # Reopen source workbook
                source_wb = openpyxl.load_workbook(self.excel_file)
                
                # Process each sheet
                for name in source_wb.sheetnames:
                    print(f"Processing sheet: {name}")
                    worksheet = workbook.add_worksheet(name)
                    source_sheet = source_wb[name]
                    
                    # Copy existing data
                    for row_idx, row in enumerate(source_sheet.rows):
                        for col_idx, cell in enumerate(row):
                            if cell.value is not None:
                                worksheet.write(row_idx, col_idx, cell.value)
                    
                    # Set column widths
                    worksheet.set_column('A:A', 40)
                    worksheet.set_column('B:B', 40)
                    worksheet.set_column('C:C', 40)
                    worksheet.set_column('D:D', 10)
                    worksheet.set_column('E:E', 20)
                    worksheet.set_column('F:F', 15)
                    worksheet.set_column('G:G', 15)
                    worksheet.set_column('H:H', 100)
                    
                    # Set row height for image
                    # worksheet.set_row(1, 400)
                    
                    if name == sheet_name:
                        # Add new image to current sheet
                        print(f"Adding new image to sheet: {name}")
                        try:
                            worksheet.insert_image(
                                'H2',
                                str(img_path),
                                {
                                    'x_offset': 10,
                                    'y_offset': 10,
                                    'x_scale': 0.7,
                                    'y_scale': 0.7,
                                    'positioning': 2,
                                    'object_position': 2
                                }
                            )
                            # Mark this sheet with its image
                            worksheet.write('H1', f'Image: {pdf_file.stem}')
                            print(f"Added new image to {name}")
                        except Exception as img_err:
                            print(f"Error inserting new image: {str(img_err)}")
                    elif name in existing_images:
                        # Restore existing image from other sheets
                        existing_img_path = self.images_dir / f"{existing_images[name]}.png"
                        if existing_img_path.exists():
                            print(f"Restoring existing image for sheet: {name}")
                            try:
                                worksheet.insert_image(
                                    'H2',
                                    str(existing_img_path),
                                    {
                                        'x_offset': 10,
                                        'y_offset': 10,
                                        'x_scale': 0.7,
                                        'y_scale': 0.7,
                                        'positioning': 2,
                                        'object_position': 2
                                    }
                                )
                                # Preserve the image marker
                                worksheet.write('H1', f'Image: {existing_images[name]}')
                                print(f"Restored existing image for {name}")
                            except Exception as img_err:
                                print(f"Error restoring existing image: {str(img_err)}")
                
                # Close source workbook
                source_wb.close()
                
                # Close and save new workbook
                workbook.close()
                
                # Replace original file
                if os.path.exists(temp_excel):
                    if os.path.exists(self.excel_file):
                        os.remove(self.excel_file)
                    os.rename(temp_excel, self.excel_file)
                    print("Excel file updated successfully")
                
            except Exception as e:
                print(f"Error processing workbook: {str(e)}")
                if workbook:
                    try:
                        workbook.close()
                    except:
                        pass
                raise
                
        except Exception as e:
            print(f"Error in image processing: {str(e)}")
            traceback.print_exc()
        finally:
            if os.path.exists(temp_excel):
                try:
                    os.remove(temp_excel)
                except Exception as e:
                    print(f"Error removing temp file: {str(e)}")

    def create_or_get_sheet(self, json_file: Path, pdf_file: Optional[Path]) -> str:
        """Create sheet if not exists or get existing sheet name with enhanced debugging."""
        print(f"\nCreating/getting sheet for {json_file.name}")
        original_sheet_name = self._clean_sheet_name(json_file.name)
        sheet_name = original_sheet_name
        
        print(f"Original sheet name: {json_file.name}")
        print(f"Cleaned sheet name: {sheet_name}")
        
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            print(f"Existing sheets in workbook: {wb.sheetnames}")
        except FileNotFoundError:
            print("Creating new workbook")
            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
                print("Removed default sheet")

        if sheet_name not in wb.sheetnames:
            print(f"Creating new sheet: {sheet_name}")
            sheet = wb.create_sheet(sheet_name)
            
            # Set up headers
            headers = ['Field', 'Gemini Value', 'Claude Value', 'Match', 'Last Updated', 'Verifier']
            for col, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col, value=header)
            
            # Store PDF path if available
            if pdf_file:
                print(f"Storing PDF path: {pdf_file}")
                sheet.cell(row=1, column=8, value=str(pdf_file))
            
            # Set column widths
            sheet.column_dimensions['A'].width = 40
            sheet.column_dimensions['B'].width = 40
            sheet.column_dimensions['C'].width = 40
            sheet.column_dimensions['D'].width = 10
            sheet.column_dimensions['E'].width = 20
            sheet.column_dimensions['F'].width = 15
            sheet.column_dimensions['G'].width = 15
            sheet.column_dimensions['H'].width = 80
            
            # Add initial data
            print("Adding initial data to sheet")
            with open(json_file, 'r', encoding='utf-8') as f:
                initial_data = json.load(f)
            
            flattened_data = self._flatten_json(initial_data)
            row = 2
            for field, value in flattened_data.items():
                sheet.cell(row=row, column=1, value=field)
                sheet.cell(row=row, column=2, value=str(value))
                row += 1
            
            print(f"Added {row-2} rows of initial data")
        else:
            print(f"Sheet {sheet_name} already exists")
        
        wb.save(self.excel_file)
        print(f"Saved workbook with sheet: {sheet_name}")
        return sheet_name

    def _flatten_json(self, json_obj: Dict) -> Dict[str, Any]:
        """Flatten nested JSON structure into dot notation."""
        flattened = {}
        
        def _flatten(obj: Any, name: str = '') -> None:
            if isinstance(obj, dict):
                for key, value in obj.items():
                    new_name = f"{name}.{key}" if name else key
                    if not self._should_ignore_field(new_name):
                        _flatten(value, new_name)
            elif isinstance(obj, list):
                for i, item in enumerate(obj):
                    new_name = f"{name}.{i}" if name else str(i)
                    if not self._should_ignore_field(new_name):
                        _flatten(item, new_name)
            else:
                if name and not self._should_ignore_field(name):
                    flattened[name] = obj
                    
        _flatten(json_obj)
        return flattened

    def _should_ignore_field(self, field_path: str) -> bool:
        """Check if field should be ignored based on patterns."""
        return any(fnmatch(field_path.lower(), pattern.lower()) 
                  for pattern in self.ignored_patterns)

    def _parse_ai_response(self, result: Any) -> Dict:
        """Parse and validate AI response with improved error handling."""
        try:
            if not result:
                print("Warning: Received empty result from AI")
                return {}
                
            if isinstance(result, str):
                # Clean and validate JSON string
                result = result.strip()
                # Find the first valid JSON object
                start = result.find('{')
                end = result.rfind('}') + 1
                if start >= 0 and end > start:
                    clean_json = result[start:end]
                    try:
                        parsed = json.loads(clean_json)
                        print("Successfully parsed JSON response")
                        return parsed
                    except json.JSONDecodeError as je:
                        print(f"JSON parsing error: {str(je)}")
                        print(f"Problematic JSON string: {clean_json[:100]}...")
                        return {}
            elif isinstance(result, dict):
                return result
                
            print(f"Unexpected result type: {type(result)}")
            return {}
            
        except Exception as e:
            print(f"Error parsing AI response: {str(e)}")
            traceback.print_exc()
            return {}
    


class BatchProcessor:
    """Orchestrates processing of multiple documents."""
    
    def __init__(self, input_folder: str, output_file: str, ignored_patterns: Optional[Set[str]] = None):
        self.input_folder = Path(input_folder)
        self.output_file = output_file
        self.images_dir = Path('invoice_images')
        
        # Default patterns if none provided
        self.ignored_patterns = ignored_patterns or {
            "supplier.supplier_phone",
            "supplier.supplier_website",
            'metadata.*',
            'technical_assessment.*',
            'metadata.processing_date',
            'signature_assessment.*',
            'line_items[*]',
            'line_items[*].*'
        }
        
        if self.images_dir.exists():
            print(f"Cleaning up existing images directory: {self.images_dir}")
            shutil.rmtree(self.images_dir)
        self.images_dir.mkdir(exist_ok=True)
        print(f"Created images directory: {self.images_dir}")



    def _find_matching_pdf(self, json_file: Path) -> Optional[Path]:
        """Find matching PDF file for a given JSON file."""
        try:
            base_numbers = json_file.stem.split('_')[0]
            print(f"Looking for PDF matching base numbers: {base_numbers}")
            
            for pdf in self.input_folder.glob('*.pdf'):
                if base_numbers in pdf.stem:
                    print(f"Found matching PDF: {pdf}")
                    return pdf
                    
            print(f"No matching PDF found for base numbers: {base_numbers}")
            return None
            
        except Exception as e:
            print(f"Error finding matching PDF for {json_file}: {str(e)}")
            return None

    def process_all_documents(self, verification_type: VerificationType = VerificationType.CLAUDE,
                            skip_image_processing: bool = False):
        """Process all documents with specified options."""
        print(f"\nStarting batch processing with {verification_type.value} verification...")
        
        doc_processor = DocumentProcessor(
            self.output_file, 
            self.images_dir,
            ignored_patterns=self.ignored_patterns
        )
        
        for json_file in self.input_folder.glob('*processed.json'):
            try:
                print(f"\nProcessing document: {json_file.name}")
                
                pdf_file = self._find_matching_pdf(json_file)
                if not pdf_file:
                    print(f"No matching PDF found for {json_file.name}, skipping...")
                    continue
                
                # Step 1: Create/get sheet
                sheet_name = doc_processor.create_or_get_sheet(json_file, pdf_file)
                
                # Step 2: Verification (if not skipped)
                if verification_type != VerificationType.SKIP:
                    if verification_type == VerificationType.CLAUDE:
                        ai_data = doc_processor.verify_with_claude(json_file, pdf_file)
                        verifier = "Claude"
                    else:  # GEMINI
                        ai_data = doc_processor.verify_with_gemini(json_file, pdf_file)
                        verifier = "Gemini"
                    
                    doc_processor.update_sheet_with_verification(sheet_name, ai_data, verifier)
                
                # Step 3: Add image (if not skipped)
                if not skip_image_processing:
                    doc_processor.add_image_to_sheet(sheet_name, pdf_file)
                
                print(f"Completed processing {json_file.name}")
                
            except Exception as e:
                print(f"Error processing {json_file.name}:")
                print(traceback.format_exc())
                continue
        
        if self.images_dir.exists():
            shutil.rmtree(self.images_dir)
        
        print("\nBatch processing completed!")


def main():
    success = load_dotenv("../.env")
    print(f"Load success: {success}")
    INPUT_DIR=os.getenv("INPUT_DIR")
    OUTPUT_DIR=os.getenv("OUTPUT_DIR")
    ERROR_DIR=os.getenv("ERROR_DIR")

 
    input_folder = OUTPUT_DIR + "/invoices"
    output_file = "processed_results.xlsx"
    
    #  Define custom patterns to ignore
    ignored_patterns = {
        "supplier.supplier_phone",
        "supplier.supplier_website",
        'metadata.*',
        'technical_assessment.*',
        'metadata.processing_date',
        'signature_assessment.*',
        'line_items[*]',
        'line_items[*].*'
    }

    # Initialize with custom patterns
    processor = BatchProcessor(
        input_folder=input_folder,
        output_file=output_file,
        ignored_patterns=ignored_patterns
    )


    
    # Example usage with different options:
    
    # Default processing (Claude verification)
    processor.process_all_documents()
    
    # Process with Gemini verification
    # processor.process_all_documents(verification_type=VerificationType.GEMINI)
    
    # Process without verification
    # processor.process_all_documents(verification_type=VerificationType.SKIP)
    
    # Process without image processing
    # processor.process_all_documents(skip_image_processing=True)
    
    # Process with Gemini and skip images
    # processor.process_all_documents(
    #     verification_type=VerificationType.GEMINI,
    #     skip_image_processing=True
    # )



if __name__ == "__main__":
    main()
    
    
    
