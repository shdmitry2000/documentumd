import openpyxl
from openpyxl.styles import PatternFill, Font
import pandas as pd
import numpy as np

def compare_values(gemini, claude):
    """Compare values considering different data types and edge cases."""
    # Handle null/None cases
    if pd.isna(gemini) and pd.isna(claude):
        return True
    if str(gemini).lower() == 'none' and pd.isna(claude):
        return True
    if pd.isna(gemini) and str(claude).lower() == 'none':
        return True
    if str(gemini).lower() == 'none' and str(claude).lower() == 'none':
        return True
    
    # If one is null and the other isn't, they don't match
    if pd.isna(gemini) or pd.isna(claude):
        return False

    # Convert to strings for comparison
    gemini_str = str(gemini).strip().lower()
    claude_str = str(claude).strip().lower()
    
    # Try numeric comparison first
    try:
        gemini_num = float(gemini_str)
        claude_num = float(claude_str)
        return abs(gemini_num - claude_num) < 0.01
    except (ValueError, TypeError):
        pass
    
    # String comparison
    return gemini_str == claude_str

def fix_worksheet_matches(ws, df_sheet):
    """Fix match calculations for a single worksheet."""
    # Calculate correct matches
    correct_matches = [
        'Yes' if compare_values(row['Gemini Value'], row['Claude Value']) else 'No'
        for _, row in df_sheet.iterrows()
    ]
    
    # Find the Match column
    header_row = 1
    match_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(header_row, col).value == 'Match':
            match_col = col
            break
    
    if match_col is None:
        raise ValueError(f"Match column not found in sheet: {ws.title}")
    
    # Define fills and fonts
    green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    white_font = Font(color='FFFFFF')  # For red background
    black_font = Font(color='000000')  # For green background
    
    # Update matches with formatting
    for row_idx, match_value in enumerate(correct_matches, start=2):  # Start from row 2 (after header)
        cell = ws.cell(row_idx, match_col)
        cell.value = match_value
        
        if match_value == 'Yes':
            cell.fill = green_fill
            cell.font = black_font
        else:
            cell.fill = red_fill
            cell.font = white_font
    
    return len(df_sheet), sum(1 for m in correct_matches if m == 'Yes'), sum(1 for m in correct_matches if m == 'No')

def fix_excel_matches_all_sheets(file_path):
    """Fix match calculations in all sheets of the Excel file while preserving formatting."""
    # Read all sheets using pandas
    xl = pd.ExcelFile(file_path)
    sheet_names = xl.sheet_names
    
    # Now use openpyxl to update the file while preserving formatting
    wb = openpyxl.load_workbook(file_path)
    
    total_summary = {
        'total_rows': 0,
        'total_matches': 0,
        'total_non_matches': 0
    }
    
    print("\nProcessing sheets:")
    print("-" * 50)
    
    for sheet_name in sheet_names:
        print(f"\nProcessing sheet: {sheet_name}")
        
        # Get the worksheet and corresponding pandas DataFrame
        ws = wb[sheet_name]
        df_sheet = pd.read_excel(file_path, sheet_name=sheet_name)
        
        try:
            # Process the sheet
            rows, matches, non_matches = fix_worksheet_matches(ws, df_sheet)
            
            # Update summary
            total_summary['total_rows'] += rows
            total_summary['total_matches'] += matches
            total_summary['total_non_matches'] += non_matches
            
            print(f"  Rows processed: {rows}")
            print(f"  Matches: {matches}")
            print(f"  Non-matches: {non_matches}")
            
        except Exception as e:
            print(f"  Error processing sheet {sheet_name}: {str(e)}")
    
    # Save the workbook
    wb.save(file_path)
    
    # Print total summary
    print("\nTotal Summary:")
    print("-" * 50)
    print(f"Total rows processed: {total_summary['total_rows']}")
    print(f"Total matches: {total_summary['total_matches']}")
    print(f"Total non-matches: {total_summary['total_non_matches']}")

if __name__ == "__main__":
    file_path = "processed_all_results.xlsx"
    
    try:
        fix_excel_matches_all_sheets(file_path)
        print(f"\nExcel file updated successfully!")
    except Exception as e:
        print(f"Error: {str(e)}")